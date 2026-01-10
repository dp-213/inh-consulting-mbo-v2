from __future__ import annotations

from io import BytesIO
from typing import Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    Workbook = None  # type: ignore[assignment]
    Alignment = Border = Font = PatternFill = Side = get_column_letter = None  # type: ignore[assignment]
    _OPENPYXL_IMPORT_ERROR = exc
else:
    _OPENPYXL_IMPORT_ERROR = None

from model.run_model import ModelResult
from state.assumptions import Assumptions

YEAR_LABELS = [f"Year {i}" for i in range(5)]


class ExcelExportError(RuntimeError):
    pass


def export_ic_excel(
    assumptions: Assumptions,
    result: ModelResult,
    case_name: str,
) -> bytes:
    if Workbook is None:
        raise ImportError("openpyxl is required for the Excel export.") from _OPENPYXL_IMPORT_ERROR
    workbook = Workbook()
    workbook.remove(workbook.active)

    styles = _styles()

    assumptions_map = _build_assumptions_sheet(
        workbook.create_sheet("Assumptions"),
        assumptions,
        case_name,
        styles,
    )
    revenue_map = _build_revenue_sheet(
        workbook.create_sheet("Revenue Model"),
        assumptions_map,
        assumptions.scenario,
        styles,
    )
    cost_map = _build_cost_sheet(
        workbook.create_sheet("Cost Model"),
        assumptions_map,
        revenue_map,
        styles,
    )
    cashflow_map = _build_cashflow_sheet(
        workbook.create_sheet("Cashflow & Liquidity"),
        assumptions_map,
        revenue_map,
        cost_map,
        styles,
    )
    debt_map = _build_debt_sheet(
        workbook.create_sheet("Financing & Debt"),
        assumptions_map,
        cashflow_map,
        styles,
    )
    pnl_map = _build_pnl_sheet(
        workbook.create_sheet("Operating Model (P&L)"),
        assumptions_map,
        revenue_map,
        cost_map,
        cashflow_map,
        debt_map,
        styles,
    )
    balance_map = _build_balance_sheet(
        workbook.create_sheet("Balance Sheet"),
        assumptions_map,
        debt_map,
        pnl_map,
        cashflow_map,
        styles,
    )
    equity_map = _build_equity_sheet(
        workbook.create_sheet("Equity Case"),
        assumptions_map,
        pnl_map,
        balance_map,
        styles,
    )
    _build_valuation_sheet(
        workbook.create_sheet("Valuation & Purchase Price"),
        assumptions_map,
        pnl_map,
        cashflow_map,
        balance_map,
        equity_map,
        styles,
    )
    _build_overview_sheet(
        workbook.create_sheet("Overview"),
        assumptions_map,
        pnl_map,
        equity_map,
        revenue_map,
        styles,
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _styles() -> Dict[str, object]:
    thin = Side(border_style="thin", color="D1D5DB")
    return {
        "title": Font(size=14, bold=True, color="111827"),
        "section": Font(size=10, bold=True, color="374151"),
        "header": Font(size=10, bold=True, color="111827"),
        "label": Font(size=9, bold=False, color="111827"),
        "label_bold": Font(size=9, bold=True, color="111827"),
        "input": Font(size=9, color="111827"),
        "output": Font(size=9, color="111827"),
        "fill_header": PatternFill("solid", fgColor="F3F4F6"),
        "fill_section": PatternFill("solid", fgColor="E5E7EB"),
        "fill_total": PatternFill("solid", fgColor="F8FAFC"),
        "fill_year0": PatternFill("solid", fgColor="E9EDF5"),
        "align_left": Alignment(horizontal="left", vertical="center"),
        "align_right": Alignment(horizontal="right", vertical="center"),
        "align_center": Alignment(horizontal="center", vertical="center"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _build_assumptions_sheet(
    ws,
    assumptions: Assumptions,
    case_name: str,
    styles: Dict[str, object],
) -> Dict[str, List[str] | str]:
    _set_column_widths(ws, [32, 12, 16, 16, 16, 16, 16])
    ws.freeze_panes = "C6"

    _write_title(
        ws,
        f"IC Model Assumptions - {case_name}",
        1,
        7,
        styles,
    )
    ws.cell(row=2, column=1, value="Scenario").font = styles["label_bold"]
    ws.cell(row=2, column=2, value=assumptions.scenario).font = styles["label"]

    start_row = 4
    assumption_cells: Dict[str, List[str] | str] = {}
    scenario = assumptions.revenue.scenarios[assumptions.scenario]

    start_row = _write_year_section(
        ws,
        start_row,
        "Revenue Assumptions",
        [
            ("Workdays per Year", "Days", scenario.workdays_per_year, "revenue.workdays"),
            ("Utilization %", "%", scenario.utilization_rate_pct, "revenue.utilization"),
            ("Day Rate Growth (% p.a.)", "% p.a.", scenario.day_rate_growth_pct, "revenue.day_rate_growth"),
            ("Revenue Growth (% p.a.)", "% p.a.", scenario.revenue_growth_pct, "revenue.revenue_growth"),
            ("Group Capacity Share %", "%", scenario.group_capacity_share_pct, "revenue.group_share"),
            ("External Capacity Share %", "%", scenario.external_capacity_share_pct, "revenue.external_share"),
            ("Group Day Rate", "EUR", scenario.group_day_rate_eur, "revenue.group_rate"),
            ("External Day Rate", "EUR", scenario.external_day_rate_eur, "revenue.external_rate"),
            ("Guarantee %", "%", scenario.guarantee_pct_by_year, "revenue.guarantee_pct"),
        ],
        styles,
        assumption_cells,
    )
    start_row = _write_value_section(
        ws,
        start_row,
        "Revenue Baseline",
        [
            ("Reference Revenue", "EUR", scenario.reference_revenue_eur, "revenue.reference"),
        ],
        styles,
        assumption_cells,
    )
    start_row = _write_value_section(
        ws,
        start_row,
        "Cost Inflation",
        [
            ("Apply Inflation", "", "TRUE" if assumptions.cost.inflation_apply else "FALSE", "cost.inflation_apply"),
            ("Inflation Rate (% p.a.)", "% p.a.", assumptions.cost.inflation_rate_pct, "cost.inflation_rate"),
        ],
        styles,
        assumption_cells,
    )
    start_row = _write_year_section(
        ws,
        start_row,
        "Personnel Costs",
        [
            ("Consultant FTE", "FTE", [row.consultant_fte for row in assumptions.cost.personnel_by_year], "cost.consultant_fte"),
            ("Consultant Loaded Cost", "EUR", [row.consultant_loaded_cost_eur for row in assumptions.cost.personnel_by_year], "cost.consultant_loaded"),
            ("Backoffice FTE", "FTE", [row.backoffice_fte for row in assumptions.cost.personnel_by_year], "cost.backoffice_fte"),
            ("Backoffice Loaded Cost", "EUR", [row.backoffice_loaded_cost_eur for row in assumptions.cost.personnel_by_year], "cost.backoffice_loaded"),
            ("Management Cost", "EUR", [row.management_cost_eur for row in assumptions.cost.personnel_by_year], "cost.management_cost"),
        ],
        styles,
        assumption_cells,
    )
    start_row = _write_year_section(
        ws,
        start_row,
        "Fixed Overhead",
        [
            ("Advisory", "EUR", [row.advisory_eur for row in assumptions.cost.fixed_overhead_by_year], "cost.advisory"),
            ("Legal", "EUR", [row.legal_eur for row in assumptions.cost.fixed_overhead_by_year], "cost.legal"),
            ("IT & Software", "EUR", [row.it_software_eur for row in assumptions.cost.fixed_overhead_by_year], "cost.it"),
            ("Office Rent", "EUR", [row.office_rent_eur for row in assumptions.cost.fixed_overhead_by_year], "cost.office"),
            ("Services", "EUR", [row.services_eur for row in assumptions.cost.fixed_overhead_by_year], "cost.services"),
            ("Other Services", "EUR", [row.other_services_eur for row in assumptions.cost.fixed_overhead_by_year], "cost.other_services"),
        ],
        styles,
        assumption_cells,
    )
    start_row = _write_year_section(
        ws,
        start_row,
        "Variable Costs",
        [
            ("Training Type", "Type", [row.training_type for row in assumptions.cost.variable_costs_by_year], "cost.training_type"),
            ("Training Value", "EUR / %", [row.training_value for row in assumptions.cost.variable_costs_by_year], "cost.training_value"),
            ("Travel Type", "Type", [row.travel_type for row in assumptions.cost.variable_costs_by_year], "cost.travel_type"),
            ("Travel Value", "EUR / %", [row.travel_value for row in assumptions.cost.variable_costs_by_year], "cost.travel_value"),
            ("Communication Type", "Type", [row.communication_type for row in assumptions.cost.variable_costs_by_year], "cost.communication_type"),
            ("Communication Value", "EUR / %", [row.communication_value for row in assumptions.cost.variable_costs_by_year], "cost.communication_value"),
        ],
        styles,
        assumption_cells,
    )
    start_row = _write_value_section(
        ws,
        start_row,
        "Transaction & Financing",
        [
            ("Purchase Price", "EUR", assumptions.transaction_and_financing.purchase_price_eur, "financing.purchase_price"),
            ("Owner Contribution", "EUR", assumptions.transaction_and_financing.equity_contribution_eur, "financing.equity_contribution"),
            ("Senior Loan Start", "EUR", assumptions.transaction_and_financing.senior_term_loan_start_eur, "financing.loan_start"),
            ("Senior Debt Amount", "EUR", assumptions.financing.senior_debt_amount_eur, "financing.senior_debt"),
            ("Opening Loan Balance", "EUR", assumptions.financing.initial_debt_eur, "financing.initial_debt"),
            ("Interest Rate", "% p.a.", assumptions.financing.interest_rate_pct, "financing.interest_rate"),
            ("Repayment Type", "", assumptions.financing.amortization_type, "financing.amort_type"),
            ("Repayment Period (Years)", "Years", assumptions.financing.amortization_period_years, "financing.amort_period"),
            ("Grace Period (Years)", "Years", assumptions.financing.grace_period_years, "financing.grace_period"),
            ("Special Repayment Year", "Year", assumptions.financing.special_repayment_year or 0, "financing.special_year"),
            ("Special Repayment Amount", "EUR", assumptions.financing.special_repayment_amount_eur, "financing.special_amount"),
            ("Minimum DSCR", "x", assumptions.financing.minimum_dscr, "financing.minimum_dscr"),
        ],
        styles,
        assumption_cells,
    )
    start_row = _write_value_section(
        ws,
        start_row,
        "Cashflow Assumptions",
        [
            ("Tax Cash Rate", "%", assumptions.cashflow.tax_cash_rate_pct, "cashflow.tax_cash_rate"),
            ("Tax Payment Lag (Years)", "Years", assumptions.cashflow.tax_payment_lag_years, "cashflow.tax_lag"),
            ("Capex (% of Revenue)", "%", assumptions.cashflow.capex_pct_revenue, "cashflow.capex_pct"),
            ("Working Capital (% of Revenue)", "%", assumptions.cashflow.working_capital_pct_revenue, "cashflow.wc_pct"),
            ("Opening Cash Balance", "EUR", assumptions.cashflow.opening_cash_balance_eur, "cashflow.opening_cash"),
        ],
        styles,
        assumption_cells,
    )
    start_row = _write_value_section(
        ws,
        start_row,
        "Balance Sheet Assumptions",
        [
            ("Opening Equity", "EUR", assumptions.balance_sheet.opening_equity_eur, "balance.opening_equity"),
            ("Depreciation Rate", "%", assumptions.balance_sheet.depreciation_rate_pct, "balance.depr_rate"),
        ],
        styles,
        assumption_cells,
    )
    start_row = _write_value_section(
        ws,
        start_row,
        "Tax & Valuation",
        [
            ("Profit Tax Rate", "%", assumptions.tax_and_distributions.tax_rate_pct, "tax.rate"),
            ("Seller Multiple", "x", assumptions.valuation.seller_multiple, "valuation.multiple"),
            ("Discount Rate (Cost of Capital)", "%", 0.10, "valuation.discount_rate"),
            ("Transaction Costs (%)", "%", 0.01, "valuation.txn_cost_pct"),
            ("Investor Exit Year", "Year", 4, "valuation.exit_year"),
            ("Minimum Cash Balance", "EUR", 250_000.0, "valuation.min_cash"),
        ],
        styles,
        assumption_cells,
    )

    return assumption_cells


def _build_revenue_sheet(
    ws,
    assumptions_map: Dict[str, List[str] | str],
    scenario: str,
    styles: Dict[str, object],
) -> Dict[str, List[str]]:
    _set_column_widths(ws, [32, 12, 16, 16, 16, 16, 16])
    ws.freeze_panes = "C6"
    _write_title(ws, f"Revenue Model - {scenario}", 1, 7, styles)

    row = 4
    row = _write_header_row(ws, row, styles)

    row_map: Dict[str, int] = {}

    row = _write_section_label(ws, row, "Consultant Capacity", styles)
    row = _write_row_from_assumptions(ws, row, "Consultant FTE", "FTE", assumptions_map["cost.consultant_fte"], styles, row_map, "consultant_fte")
    row = _write_row_from_assumptions(ws, row, "Workdays per Year", "Days", assumptions_map["revenue.workdays"], styles, row_map, "workdays")
    row = _write_row_from_assumptions(ws, row, "Utilization %", "%", assumptions_map["revenue.utilization"], styles, row_map, "utilization")
    row = _write_formula_row(
        ws,
        row,
        "Capacity Days",
        "Days",
        row_map,
        "capacity_days",
        styles,
        lambda c: f"=C{row_map['consultant_fte']}*C{row_map['workdays']}*(C{row_map['utilization']}/100)",
    )
    row = _write_row_from_assumptions(ws, row, "Revenue Growth (% p.a.)", "% p.a.", assumptions_map["revenue.revenue_growth"], styles, row_map, "revenue_growth")
    row = _write_formula_row(
        ws,
        row,
        "Adjusted Capacity Days",
        "Days",
        row_map,
        "adjusted_capacity",
        styles,
        lambda c: f"=C{row_map['capacity_days']}*(1+C{row_map['revenue_growth']}/100)",
    )

    row = _write_section_label(ws, row + 1, "Capacity Allocation", styles)
    row = _write_row_from_assumptions(ws, row, "Group Capacity Share %", "%", assumptions_map["revenue.group_share"], styles, row_map, "group_share")
    row = _write_row_from_assumptions(ws, row, "External Capacity Share %", "%", assumptions_map["revenue.external_share"], styles, row_map, "external_share")
    row = _write_formula_row(ws, row, "Total Share", "", row_map, "total_share", styles, lambda c: f"=C{row_map['group_share']}+C{row_map['external_share']}")
    row = _write_formula_row(
        ws,
        row,
        "Group Share (Normalized)",
        "%",
        row_map,
        "group_share_norm",
        styles,
        lambda c: f"=IF(C{row_map['total_share']}=0,0,(C{row_map['group_share']}/C{row_map['total_share']})*100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "External Share (Normalized)",
        "%",
        row_map,
        "external_share_norm",
        styles,
        lambda c: f"=IF(C{row_map['total_share']}=0,0,(C{row_map['external_share']}/C{row_map['total_share']})*100)",
    )

    row = _write_section_label(ws, row + 1, "Pricing", styles)
    row = _write_row_from_assumptions(ws, row, "Group Day Rate (Base)", "EUR", assumptions_map["revenue.group_rate"], styles, row_map, "group_rate")
    row = _write_row_from_assumptions(ws, row, "External Day Rate (Base)", "EUR", assumptions_map["revenue.external_rate"], styles, row_map, "external_rate")
    row = _write_row_from_assumptions(ws, row, "Day Rate Growth (% p.a.)", "% p.a.", assumptions_map["revenue.day_rate_growth"], styles, row_map, "day_rate_growth")
    row = _write_formula_row(
        ws,
        row,
        "Group Day Rate (Indexed)",
        "EUR",
        row_map,
        "group_rate_indexed",
        styles,
        lambda c, idx=None: f"=C{row_map['group_rate']}*(1+C{row_map['day_rate_growth']}/100)^{idx}",
    )
    row = _write_formula_row(
        ws,
        row,
        "External Day Rate (Indexed)",
        "EUR",
        row_map,
        "external_rate_indexed",
        styles,
        lambda c, idx=None: f"=C{row_map['external_rate']}*(1+C{row_map['day_rate_growth']}/100)^{idx}",
    )

    row = _write_section_label(ws, row + 1, "Revenue Bridge", styles)
    row = _write_formula_row(
        ws,
        row,
        "Modeled Group Revenue",
        "EUR",
        row_map,
        "modeled_group_revenue",
        styles,
        lambda c: f"=C{row_map['adjusted_capacity']}*(C{row_map['group_share_norm']}/100)*C{row_map['group_rate_indexed']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Modeled External Revenue",
        "EUR",
        row_map,
        "modeled_external_revenue",
        styles,
        lambda c: f"=C{row_map['adjusted_capacity']}*(C{row_map['external_share_norm']}/100)*C{row_map['external_rate_indexed']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Modeled Total Revenue",
        "EUR",
        row_map,
        "modeled_total_revenue",
        styles,
        lambda c: f"=C{row_map['modeled_group_revenue']}+C{row_map['modeled_external_revenue']}",
    )
    row = _write_row_from_assumptions(ws, row, "Reference Revenue", "EUR", [assumptions_map["revenue.reference"]] * 5, styles, row_map, "reference_revenue")
    row = _write_row_from_assumptions(ws, row, "Guarantee %", "%", assumptions_map["revenue.guarantee_pct"], styles, row_map, "guarantee_pct")
    row = _write_formula_row(
        ws,
        row,
        "Guaranteed Floor",
        "EUR",
        row_map,
        "guaranteed_floor",
        styles,
        lambda c: f"=C{row_map['reference_revenue']}*(C{row_map['guarantee_pct']}/100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Guaranteed Group Revenue",
        "EUR",
        row_map,
        "guaranteed_group_revenue",
        styles,
        lambda c: f"=MAX(C{row_map['modeled_group_revenue']},C{row_map['guaranteed_floor']})",
    )
    row = _write_formula_row(
        ws,
        row,
        "Final Total Revenue",
        "EUR",
        row_map,
        "final_total_revenue",
        styles,
        lambda c: f"=C{row_map['guaranteed_group_revenue']}+C{row_map['modeled_external_revenue']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Share Guaranteed",
        "%",
        row_map,
        "share_guaranteed",
        styles,
        lambda c: f"=IF(C{row_map['final_total_revenue']}=0,0,(C{row_map['guaranteed_group_revenue']}/C{row_map['final_total_revenue']})*100)",
    )

    total_row = row_map["final_total_revenue"]
    _apply_total_style(ws, total_row, styles, 7)

    return {
        "final_total_revenue": _row_cells("Revenue Model", row_map["final_total_revenue"]),
        "modeled_group_revenue": _row_cells("Revenue Model", row_map["modeled_group_revenue"]),
        "modeled_external_revenue": _row_cells("Revenue Model", row_map["modeled_external_revenue"]),
    }


def _build_cost_sheet(
    ws,
    assumptions_map: Dict[str, List[str] | str],
    revenue_map: Dict[str, List[str]],
    styles: Dict[str, object],
) -> Dict[str, List[str]]:
    _set_column_widths(ws, [34, 12, 16, 16, 16, 16, 16])
    ws.freeze_panes = "C6"
    _write_title(ws, "Cost Model", 1, 7, styles)

    row = 4
    row = _write_header_row(ws, row, styles)

    row_map: Dict[str, int] = {}

    row = _write_section_label(ws, row, "Inflation", styles)
    row = _write_row_from_assumptions(ws, row, "Apply Inflation", "", [assumptions_map["cost.inflation_apply"]] * 5, styles, row_map, "inflation_apply")
    row = _write_row_from_assumptions(ws, row, "Inflation Rate (% p.a.)", "% p.a.", [assumptions_map["cost.inflation_rate"]] * 5, styles, row_map, "inflation_rate")
    row = _write_formula_row(
        ws,
        row,
        "Inflation Factor",
        "",
        row_map,
        "inflation_factor",
        styles,
        lambda c, idx=None: f"=IF(UPPER(C{row_map['inflation_apply']})=\"TRUE\",(1+C{row_map['inflation_rate']}/100)^{idx},1)",
    )

    row = _write_section_label(ws, row + 1, "Personnel Costs", styles)
    row = _write_row_from_assumptions(ws, row, "Consultant FTE", "FTE", assumptions_map["cost.consultant_fte"], styles, row_map, "consultant_fte")
    row = _write_row_from_assumptions(ws, row, "Consultant Loaded Cost", "EUR", assumptions_map["cost.consultant_loaded"], styles, row_map, "consultant_loaded")
    row = _write_formula_row(
        ws,
        row,
        "Consultant Cost",
        "EUR",
        row_map,
        "consultant_cost",
        styles,
        lambda c: f"=C{row_map['consultant_fte']}*C{row_map['consultant_loaded']}*C{row_map['inflation_factor']}",
    )
    row = _write_row_from_assumptions(ws, row, "Backoffice FTE", "FTE", assumptions_map["cost.backoffice_fte"], styles, row_map, "backoffice_fte")
    row = _write_row_from_assumptions(ws, row, "Backoffice Loaded Cost", "EUR", assumptions_map["cost.backoffice_loaded"], styles, row_map, "backoffice_loaded")
    row = _write_formula_row(
        ws,
        row,
        "Backoffice Cost",
        "EUR",
        row_map,
        "backoffice_cost",
        styles,
        lambda c: f"=C{row_map['backoffice_fte']}*C{row_map['backoffice_loaded']}*C{row_map['inflation_factor']}",
    )
    row = _write_row_from_assumptions(ws, row, "Management Cost", "EUR", assumptions_map["cost.management_cost"], styles, row_map, "management_cost")
    row = _write_formula_row(
        ws,
        row,
        "Management Cost (Inflated)",
        "EUR",
        row_map,
        "management_cost_inflated",
        styles,
        lambda c: f"=C{row_map['management_cost']}*C{row_map['inflation_factor']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Personnel Costs",
        "EUR",
        row_map,
        "total_personnel",
        styles,
        lambda c: f"=C{row_map['consultant_cost']}+C{row_map['backoffice_cost']}+C{row_map['management_cost_inflated']}",
    )
    _apply_total_style(ws, row_map["total_personnel"], styles, 7)

    row = _write_section_label(ws, row + 1, "Fixed Overhead", styles)
    row = _write_row_from_assumptions(ws, row, "Advisory", "EUR", assumptions_map["cost.advisory"], styles, row_map, "advisory")
    row = _write_row_from_assumptions(ws, row, "Legal", "EUR", assumptions_map["cost.legal"], styles, row_map, "legal")
    row = _write_row_from_assumptions(ws, row, "IT & Software", "EUR", assumptions_map["cost.it"], styles, row_map, "it_software")
    row = _write_row_from_assumptions(ws, row, "Office Rent", "EUR", assumptions_map["cost.office"], styles, row_map, "office_rent")
    row = _write_row_from_assumptions(ws, row, "Services", "EUR", assumptions_map["cost.services"], styles, row_map, "services")
    row = _write_row_from_assumptions(ws, row, "Other Services", "EUR", assumptions_map["cost.other_services"], styles, row_map, "other_services")
    row = _write_formula_row(
        ws,
        row,
        "Fixed Overhead Total",
        "EUR",
        row_map,
        "fixed_total",
        styles,
        lambda c: (
            f"=(C{row_map['advisory']}+C{row_map['legal']}+C{row_map['it_software']}"
            f"+C{row_map['office_rent']}+C{row_map['services']}+C{row_map['other_services']})"
            f"*C{row_map['inflation_factor']}"
        ),
    )

    row = _write_section_label(ws, row + 1, "Variable Costs", styles)
    row = _write_row_from_assumptions(ws, row, "Training Type", "Type", assumptions_map["cost.training_type"], styles, row_map, "training_type")
    row = _write_row_from_assumptions(ws, row, "Training Value", "EUR / %", assumptions_map["cost.training_value"], styles, row_map, "training_value")
    row = _write_formula_row(
        ws,
        row,
        "Training Cost",
        "EUR",
        row_map,
        "training_cost",
        styles,
        lambda c, idx=None: f"=IF(UPPER(C{row_map['training_type']})=\"%\",{revenue_map['final_total_revenue'][idx]}*(C{row_map['training_value']}/100),C{row_map['training_value']}*C{row_map['inflation_factor']})",
    )
    row = _write_row_from_assumptions(ws, row, "Travel Type", "Type", assumptions_map["cost.travel_type"], styles, row_map, "travel_type")
    row = _write_row_from_assumptions(ws, row, "Travel Value", "EUR / %", assumptions_map["cost.travel_value"], styles, row_map, "travel_value")
    row = _write_formula_row(
        ws,
        row,
        "Travel Cost",
        "EUR",
        row_map,
        "travel_cost",
        styles,
        lambda c, idx=None: f"=IF(UPPER(C{row_map['travel_type']})=\"%\",{revenue_map['final_total_revenue'][idx]}*(C{row_map['travel_value']}/100),C{row_map['travel_value']}*C{row_map['inflation_factor']})",
    )
    row = _write_row_from_assumptions(ws, row, "Communication Type", "Type", assumptions_map["cost.communication_type"], styles, row_map, "communication_type")
    row = _write_row_from_assumptions(ws, row, "Communication Value", "EUR / %", assumptions_map["cost.communication_value"], styles, row_map, "communication_value")
    row = _write_formula_row(
        ws,
        row,
        "Communication Cost",
        "EUR",
        row_map,
        "communication_cost",
        styles,
        lambda c, idx=None: f"=IF(UPPER(C{row_map['communication_type']})=\"%\",{revenue_map['final_total_revenue'][idx]}*(C{row_map['communication_value']}/100),C{row_map['communication_value']}*C{row_map['inflation_factor']})",
    )
    row = _write_formula_row(
        ws,
        row,
        "Variable Costs Total",
        "EUR",
        row_map,
        "variable_total",
        styles,
        lambda c: f"=C{row_map['training_cost']}+C{row_map['travel_cost']}+C{row_map['communication_cost']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Operating Expenses",
        "EUR",
        row_map,
        "operating_expenses",
        styles,
        lambda c: f"=C{row_map['fixed_total']}+C{row_map['variable_total']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Operating Costs",
        "EUR",
        row_map,
        "total_operating_costs",
        styles,
        lambda c: f"=C{row_map['total_personnel']}+C{row_map['operating_expenses']}",
    )
    _apply_total_style(ws, row_map["operating_expenses"], styles, 7)
    _apply_total_style(ws, row_map["total_operating_costs"], styles, 7)

    return {
        "total_personnel": _row_cells("Cost Model", row_map["total_personnel"]),
        "operating_expenses": _row_cells("Cost Model", row_map["operating_expenses"]),
    }


def _build_cashflow_sheet(
    ws,
    assumptions_map: Dict[str, List[str] | str],
    revenue_map: Dict[str, List[str]],
    cost_map: Dict[str, List[str]],
    styles: Dict[str, object],
) -> Dict[str, List[str]]:
    _set_column_widths(ws, [32, 12, 16, 16, 16, 16, 16])
    ws.freeze_panes = "C6"
    _write_title(ws, "Cashflow & Liquidity", 1, 7, styles)

    row = 4
    row = _write_header_row(ws, row, styles)
    row_map: Dict[str, int] = {}

    row = _write_section_label(ws, row, "Operating Cashflow", styles)
    row = _write_formula_row(
        ws,
        row,
        "Operating Profit (EBITDA)",
        "EUR",
        row_map,
        "ebitda",
        styles,
        lambda c, idx=None: f"={revenue_map['final_total_revenue'][idx]}-{cost_map['total_personnel'][idx]}-{cost_map['operating_expenses'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Taxes Paid",
        "EUR",
        row_map,
        "taxes_paid",
        styles,
        lambda c, idx=None: (
            f"=IF({assumptions_map['cashflow.tax_lag']}=0,C{row_map['taxes_due']},"
            f"IF({assumptions_map['cashflow.tax_lag']}=1,IF({idx}=0,0,INDEX(C{row_map['taxes_due']}:G{row_map['taxes_due']},{idx})),0))"
        ),
    )
    row = _write_formula_row(
        ws,
        row,
        "Working Capital Change",
        "EUR",
        row_map,
        "working_capital_change",
        styles,
        lambda c, idx=None: (
            f"=IF({idx}=0,C{row_map['working_capital_balance']},"
            f"C{row_map['working_capital_balance']}-INDEX(C{row_map['working_capital_balance']}:G{row_map['working_capital_balance']},{idx}))"
        ),
    )
    row = _write_formula_row(
        ws,
        row,
        "Operating Cashflow",
        "EUR",
        row_map,
        "operating_cf",
        styles,
        lambda c: f"=C{row_map['ebitda']}-C{row_map['taxes_paid']}-C{row_map['working_capital_change']}",
    )

    row = _write_section_label(ws, row + 1, "Investing Cashflow", styles)
    row = _write_formula_row(
        ws,
        row,
        "Capex",
        "EUR",
        row_map,
        "capex",
        styles,
        lambda c, idx=None: f"={revenue_map['final_total_revenue'][idx]}*({assumptions_map['cashflow.capex_pct']}/100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Acquisition Outflow",
        "EUR",
        row_map,
        "acquisition_outflow",
        styles,
        lambda c, idx=None: f"=IF({idx}=0,-{assumptions_map['financing.purchase_price']},0)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Free Cashflow",
        "EUR",
        row_map,
        "free_cashflow",
        styles,
        lambda c: f"=C{row_map['operating_cf']}-C{row_map['capex']}+C{row_map['acquisition_outflow']}",
    )

    row = _write_section_label(ws, row + 1, "Financing Cashflow", styles)
    row = _write_formula_row(
        ws,
        row,
        "Debt Drawdowns",
        "EUR",
        row_map,
        "debt_drawdown",
        styles,
        lambda c, idx=None: f"=IF({idx}=0,{assumptions_map['financing.senior_debt']},0)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Equity Injection",
        "EUR",
        row_map,
        "equity_injection",
        styles,
        lambda c, idx=None: f"=IF({idx}=0,{assumptions_map['financing.equity_contribution']},0)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Interest Paid",
        "EUR",
        row_map,
        "interest_paid",
        styles,
        lambda c, idx=None: f"=C{row_map['opening_debt']}*({assumptions_map['financing.interest_rate']}/100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Debt Repayment",
        "EUR",
        row_map,
        "debt_repayment",
        styles,
        lambda c: f"=C{row_map['total_repayment']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Net Cashflow",
        "EUR",
        row_map,
        "net_cashflow",
        styles,
        lambda c, idx=None: (
            f"=C{row_map['free_cashflow']}+IF({idx}=0,C{row_map['debt_drawdown']}+C{row_map['equity_injection']}-C{row_map['interest_paid']}-C{row_map['debt_repayment']},"
            f"-(C{row_map['interest_paid']}+C{row_map['debt_repayment']}))"
        ),
    )

    row = _write_section_label(ws, row + 1, "Liquidity", styles)
    row = _write_formula_row(
        ws,
        row,
        "Opening Cash",
        "EUR",
        row_map,
        "opening_cash",
        styles,
        lambda c, idx=None: f"=IF({idx}=0,{assumptions_map['cashflow.opening_cash']},INDEX(C{row_map['closing_cash']}:G{row_map['closing_cash']},{idx}))",
    )
    row = _write_formula_row(
        ws,
        row,
        "Closing Cash",
        "EUR",
        row_map,
        "closing_cash",
        styles,
        lambda c: f"=C{row_map['opening_cash']}+C{row_map['net_cashflow']}",
    )

    row = _write_section_label(ws, row + 2, "Debt Schedule (Support)", styles)
    ws.row_dimensions[row].hidden = True
    row = _write_formula_row(
        ws,
        row,
        "Opening Debt",
        "EUR",
        row_map,
        "opening_debt",
        styles,
        lambda c, idx=None: f"=IF({idx}=0,{assumptions_map['financing.senior_debt']},INDEX(C{row_map['closing_debt']}:G{row_map['closing_debt']},{idx}))",
    )
    ws.row_dimensions[row_map["opening_debt"]].hidden = True
    row = _write_formula_row(
        ws,
        row,
        "Scheduled Repayment",
        "EUR",
        row_map,
        "scheduled_repayment",
        styles,
        lambda c, idx=None: (
            f"=IF(UPPER({assumptions_map['financing.amort_type']})=\"BULLET\","
            f"IF({idx}={assumptions_map['financing.amort_period']}-1,C{row_map['opening_debt']},0),"
            f"IF({idx}<{assumptions_map['financing.grace_period']},0,"
            f"IF({idx}<{assumptions_map['financing.amort_period']},C{row_map['opening_debt']}/{assumptions_map['financing.amort_period']},0)))"
        ),
    )
    ws.row_dimensions[row_map["scheduled_repayment"]].hidden = True
    row = _write_formula_row(
        ws,
        row,
        "Special Repayment",
        "EUR",
        row_map,
        "special_repayment",
        styles,
        lambda c, idx=None: f"=IF({idx}={assumptions_map['financing.special_year']},{assumptions_map['financing.special_amount']},0)",
    )
    ws.row_dimensions[row_map["special_repayment"]].hidden = True
    row = _write_formula_row(
        ws,
        row,
        "Total Repayment",
        "EUR",
        row_map,
        "total_repayment",
        styles,
        lambda c: f"=MIN(C{row_map['opening_debt']},C{row_map['scheduled_repayment']}+C{row_map['special_repayment']})",
    )
    ws.row_dimensions[row_map["total_repayment"]].hidden = True
    row = _write_formula_row(
        ws,
        row,
        "Closing Debt",
        "EUR",
        row_map,
        "closing_debt",
        styles,
        lambda c: f"=MAX(C{row_map['opening_debt']}-C{row_map['total_repayment']},0)",
    )
    ws.row_dimensions[row_map["closing_debt"]].hidden = True

    row = _write_section_label(ws, row + 2, "Supporting Calculations", styles)
    ws.row_dimensions[row].hidden = True
    row = _write_formula_row(
        ws,
        row,
        "Working Capital Balance",
        "EUR",
        row_map,
        "working_capital_balance",
        styles,
        lambda c, idx=None: f"={revenue_map['final_total_revenue'][idx]}*({assumptions_map['cashflow.wc_pct']}/100)",
    )
    ws.row_dimensions[row_map["working_capital_balance"]].hidden = True
    row = _write_formula_row(
        ws,
        row,
        "Depreciation",
        "EUR",
        row_map,
        "depreciation",
        styles,
        lambda c, idx=None: (
            f"=(({assumptions_map['balance.depr_rate']}/100)*(IF({idx}=0,0,INDEX(C{row_map['fixed_assets']}:G{row_map['fixed_assets']},{idx}))"
            f"+C{row_map['capex']}))"
        ),
    )
    ws.row_dimensions[row_map["depreciation"]].hidden = True
    row = _write_formula_row(
        ws,
        row,
        "Fixed Assets",
        "EUR",
        row_map,
        "fixed_assets",
        styles,
        lambda c, idx=None: (
            f"=MAX(IF({idx}=0,0,INDEX(C{row_map['fixed_assets']}:G{row_map['fixed_assets']},{idx}))"
            f"+C{row_map['capex']}-C{row_map['depreciation']},0)"
        ),
    )
    ws.row_dimensions[row_map["fixed_assets"]].hidden = True
    row = _write_formula_row(
        ws,
        row,
        "Taxes Due",
        "EUR",
        row_map,
        "taxes_due",
        styles,
        lambda c: f"=IF((C{row_map['ebitda']}-C{row_map['depreciation']}-C{row_map['interest_paid']})>0,(C{row_map['ebitda']}-C{row_map['depreciation']}-C{row_map['interest_paid']})*({assumptions_map['cashflow.tax_cash_rate']}/100),0)",
    )
    ws.row_dimensions[row_map["taxes_due"]].hidden = True

    _apply_total_style(ws, row_map["operating_cf"], styles, 7)
    _apply_total_style(ws, row_map["free_cashflow"], styles, 7)
    _apply_total_style(ws, row_map["net_cashflow"], styles, 7)
    _apply_total_style(ws, row_map["closing_cash"], styles, 7)

    return {
        "operating_cf": _row_cells("Cashflow & Liquidity", row_map["operating_cf"]),
        "free_cashflow": _row_cells("Cashflow & Liquidity", row_map["free_cashflow"]),
        "net_cashflow": _row_cells("Cashflow & Liquidity", row_map["net_cashflow"]),
        "closing_cash": _row_cells("Cashflow & Liquidity", row_map["closing_cash"]),
        "opening_cash": _row_cells("Cashflow & Liquidity", row_map["opening_cash"]),
        "working_capital_balance": _row_cells("Cashflow & Liquidity", row_map["working_capital_balance"]),
        "taxes_due": _row_cells("Cashflow & Liquidity", row_map["taxes_due"]),
        "taxes_paid": _row_cells("Cashflow & Liquidity", row_map["taxes_paid"]),
        "fixed_assets": _row_cells("Cashflow & Liquidity", row_map["fixed_assets"]),
        "depreciation": _row_cells("Cashflow & Liquidity", row_map["depreciation"]),
        "capex": _row_cells("Cashflow & Liquidity", row_map["capex"]),
        "interest_paid": _row_cells("Cashflow & Liquidity", row_map["interest_paid"]),
        "total_repayment": _row_cells("Cashflow & Liquidity", row_map["total_repayment"]),
        "opening_debt": _row_cells("Cashflow & Liquidity", row_map["opening_debt"]),
        "closing_debt": _row_cells("Cashflow & Liquidity", row_map["closing_debt"]),
    }


def _build_debt_sheet(
    ws,
    assumptions_map: Dict[str, List[str] | str],
    cashflow_map: Dict[str, List[str]],
    styles: Dict[str, object],
) -> Dict[str, List[str]]:
    _set_column_widths(ws, [32, 14, 16, 16, 16, 16, 16])
    ws.freeze_panes = "C6"
    _write_title(ws, "Financing & Debt", 1, 7, styles)

    row = 4
    row = _write_header_row(ws, row, styles)
    row_map: Dict[str, int] = {}

    row = _write_section_label(ws, row, "Bank View", styles)
    row = _write_formula_row(
        ws,
        row,
        "Operating Profit (EBITDA)",
        "EUR",
        row_map,
        "ebitda",
        styles,
        lambda c, idx=None: f"={cashflow_map['operating_cf'][idx]}+{cashflow_map['taxes_paid'][idx]}+{cashflow_map['working_capital_balance'][idx]}-{cashflow_map['working_capital_balance'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Cash Taxes",
        "EUR",
        row_map,
        "taxes_paid",
        styles,
        lambda c, idx=None: f"={cashflow_map['taxes_paid'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Capex (Maintenance)",
        "EUR",
        row_map,
        "capex",
        styles,
        lambda c, idx=None: f"={cashflow_map['capex'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Working Capital Change",
        "EUR",
        row_map,
        "working_capital_change",
        styles,
        lambda c, idx=None: f"=0",
    )
    row = _write_formula_row(
        ws,
        row,
        "Cash Available for Debt Service",
        "EUR",
        row_map,
        "cfads",
        styles,
        lambda c: f"=C{row_map['ebitda']}-C{row_map['taxes_paid']}-C{row_map['capex']}-C{row_map['working_capital_change']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Interest Expense",
        "EUR",
        row_map,
        "interest_expense",
        styles,
        lambda c, idx=None: f"={cashflow_map['interest_paid'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Scheduled Repayment",
        "EUR",
        row_map,
        "scheduled_repayment",
        styles,
        lambda c, idx=None: f"={cashflow_map['total_repayment'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Debt Service",
        "EUR",
        row_map,
        "debt_service",
        styles,
        lambda c: f"=C{row_map['interest_expense']}+C{row_map['scheduled_repayment']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Debt Service Coverage",
        "x",
        row_map,
        "dscr",
        styles,
        lambda c: f"=IF(C{row_map['debt_service']}=0,0,C{row_map['cfads']}/C{row_map['debt_service']})",
    )
    row = _write_formula_row(
        ws,
        row,
        "Minimum Required Coverage",
        "x",
        row_map,
        "minimum_dscr",
        styles,
        lambda c: f"={assumptions_map['financing.minimum_dscr']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Covenant Breach",
        "",
        row_map,
        "covenant_breach",
        styles,
        lambda c: f"=IF(C{row_map['dscr']}<C{row_map['minimum_dscr']},\"YES\",\"NO\")",
    )

    _apply_total_style(ws, row_map["cfads"], styles, 7)
    _apply_total_style(ws, row_map["debt_service"], styles, 7)

    return {
        "interest_expense": _row_cells("Financing & Debt", row_map["interest_expense"]),
        "total_repayment": _row_cells("Financing & Debt", row_map["scheduled_repayment"]),
        "opening_debt": cashflow_map["opening_debt"],
        "closing_debt": cashflow_map["closing_debt"],
    }


def _build_pnl_sheet(
    ws,
    assumptions_map: Dict[str, List[str] | str],
    revenue_map: Dict[str, List[str]],
    cost_map: Dict[str, List[str]],
    cashflow_map: Dict[str, List[str]],
    debt_map: Dict[str, List[str]],
    styles: Dict[str, object],
) -> Dict[str, List[str]]:
    _set_column_widths(ws, [32, 12, 16, 16, 16, 16, 16])
    ws.freeze_panes = "C6"
    _write_title(ws, "Operating Model (P&L)", 1, 7, styles)

    row = 4
    row = _write_header_row(ws, row, styles)
    row_map: Dict[str, int] = {}

    row = _write_section_label(ws, row, "Revenue", styles)
    row = _write_formula_row(
        ws,
        row,
        "Total Revenue",
        "EUR",
        row_map,
        "total_revenue",
        styles,
        lambda c, idx=None: f"={revenue_map['final_total_revenue'][idx]}",
    )
    row = _write_section_label(ws, row + 1, "Costs", styles)
    row = _write_formula_row(
        ws,
        row,
        "Total Personnel Costs",
        "EUR",
        row_map,
        "personnel_costs",
        styles,
        lambda c, idx=None: f"={cost_map['total_personnel'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Operating Expenses",
        "EUR",
        row_map,
        "operating_expenses",
        styles,
        lambda c, idx=None: f"={cost_map['operating_expenses'][idx]}",
    )
    row = _write_section_label(ws, row + 1, "Profitability", styles)
    row = _write_formula_row(
        ws,
        row,
        "EBITDA",
        "EUR",
        row_map,
        "ebitda",
        styles,
        lambda c: f"=C{row_map['total_revenue']}-C{row_map['personnel_costs']}-C{row_map['operating_expenses']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Depreciation",
        "EUR",
        row_map,
        "depreciation",
        styles,
        lambda c, idx=None: f"={cashflow_map['depreciation'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "EBIT",
        "EUR",
        row_map,
        "ebit",
        styles,
        lambda c: f"=C{row_map['ebitda']}-C{row_map['depreciation']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Interest Expense",
        "EUR",
        row_map,
        "interest_expense",
        styles,
        lambda c, idx=None: f"={debt_map['interest_expense'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "EBT",
        "EUR",
        row_map,
        "ebt",
        styles,
        lambda c: f"=C{row_map['ebit']}-C{row_map['interest_expense']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Taxes",
        "EUR",
        row_map,
        "taxes",
        styles,
        lambda c: f"=IF(C{row_map['ebt']}>0,C{row_map['ebt']}*({assumptions_map['cashflow.tax_cash_rate']}/100),0)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Net Income",
        "EUR",
        row_map,
        "net_income",
        styles,
        lambda c: f"=C{row_map['ebt']}-C{row_map['taxes']}",
    )

    row = _write_section_label(ws, row + 1, "KPIs", styles)
    row = _write_formula_row(
        ws,
        row,
        "Revenue per Consultant",
        "EUR",
        row_map,
        "revenue_per_consultant",
        styles,
        lambda c, idx=None: f"=IF({assumptions_map['cost.consultant_fte'][idx]}=0,0,C{row_map['total_revenue']}/{assumptions_map['cost.consultant_fte'][idx]})",
    )
    row = _write_formula_row(
        ws,
        row,
        "EBITDA Margin",
        "%",
        row_map,
        "ebitda_margin",
        styles,
        lambda c: f"=IF(C{row_map['total_revenue']}=0,0,(C{row_map['ebitda']}/C{row_map['total_revenue']})*100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "EBIT Margin",
        "%",
        row_map,
        "ebit_margin",
        styles,
        lambda c: f"=IF(C{row_map['total_revenue']}=0,0,(C{row_map['ebit']}/C{row_map['total_revenue']})*100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Personnel Cost Ratio",
        "%",
        row_map,
        "personnel_ratio",
        styles,
        lambda c: f"=IF(C{row_map['total_revenue']}=0,0,(C{row_map['personnel_costs']}/C{row_map['total_revenue']})*100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Net Margin",
        "%",
        row_map,
        "net_margin",
        styles,
        lambda c: f"=IF(C{row_map['total_revenue']}=0,0,(C{row_map['net_income']}/C{row_map['total_revenue']})*100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Opex Ratio",
        "%",
        row_map,
        "opex_ratio",
        styles,
        lambda c: f"=IF(C{row_map['total_revenue']}=0,0,(C{row_map['operating_expenses']}/C{row_map['total_revenue']})*100)",
    )

    for key in ["total_revenue", "personnel_costs", "operating_expenses", "ebitda", "ebit", "net_income"]:
        _apply_total_style(ws, row_map[key], styles, 7)

    return {
        "ebitda": _row_cells("Operating Model (P&L)", row_map["ebitda"]),
        "ebit": _row_cells("Operating Model (P&L)", row_map["ebit"]),
        "taxes": _row_cells("Operating Model (P&L)", row_map["taxes"]),
        "net_income": _row_cells("Operating Model (P&L)", row_map["net_income"]),
        "depreciation": _row_cells("Operating Model (P&L)", row_map["depreciation"]),
    }


def _build_balance_sheet(
    ws,
    assumptions_map: Dict[str, List[str] | str],
    debt_map: Dict[str, List[str]],
    pnl_map: Dict[str, List[str]],
    cashflow_map: Dict[str, List[str]],
    styles: Dict[str, object],
) -> Dict[str, List[str]]:
    _set_column_widths(ws, [32, 12, 16, 16, 16, 16, 16])
    ws.freeze_panes = "C6"
    _write_title(ws, "Balance Sheet", 1, 7, styles)

    row = 4
    row = _write_header_row(ws, row, styles)
    row_map: Dict[str, int] = {}

    row = _write_section_label(ws, row, "Assets", styles)
    row = _write_formula_row(
        ws,
        row,
        "Cash",
        "EUR",
        row_map,
        "cash",
        styles,
        lambda c, idx=None: f"={cashflow_map['closing_cash'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Fixed Assets (Net)",
        "EUR",
        row_map,
        "fixed_assets",
        styles,
        lambda c, idx=None: f"={cashflow_map['fixed_assets'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Acquisition Intangible",
        "EUR",
        row_map,
        "acquisition_intangible",
        styles,
        lambda c: f"={assumptions_map['financing.purchase_price']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Working Capital",
        "EUR",
        row_map,
        "working_capital",
        styles,
        lambda c, idx=None: f"={cashflow_map['working_capital_balance'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Assets",
        "EUR",
        row_map,
        "total_assets",
        styles,
        lambda c: f"=C{row_map['cash']}+C{row_map['fixed_assets']}+C{row_map['acquisition_intangible']}+C{row_map['working_capital']}",
    )

    row = _write_section_label(ws, row + 1, "Liabilities", styles)
    row = _write_formula_row(
        ws,
        row,
        "Financial Debt",
        "EUR",
        row_map,
        "financial_debt",
        styles,
        lambda c, idx=None: f"={debt_map['closing_debt'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Tax Payable",
        "EUR",
        row_map,
        "tax_payable",
        styles,
        lambda c, idx=None: (
            f"=IF({idx}=0,0,INDEX(C{row_map['tax_payable']}:G{row_map['tax_payable']},{idx}))"
            f"+{cashflow_map['taxes_due'][idx]}-{cashflow_map['taxes_paid'][idx]}"
        ),
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Liabilities",
        "EUR",
        row_map,
        "total_liabilities",
        styles,
        lambda c: f"=C{row_map['financial_debt']}+C{row_map['tax_payable']}",
    )

    row = _write_section_label(ws, row + 1, "Equity", styles)
    row = _write_formula_row(
        ws,
        row,
        "Equity at Start of Year",
        "EUR",
        row_map,
        "equity_start",
        styles,
        lambda c, idx=None: f"=IF({idx}=0,{assumptions_map['balance.opening_equity']},INDEX(C{row_map['equity_end']}:G{row_map['equity_end']},{idx}))",
    )
    row = _write_formula_row(
        ws,
        row,
        "Net Income",
        "EUR",
        row_map,
        "net_income",
        styles,
        lambda c, idx=None: f"={pnl_map['net_income'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Dividends",
        "EUR",
        row_map,
        "dividends",
        styles,
        lambda c: "=0",
    )
    row = _write_formula_row(
        ws,
        row,
        "Equity Injections",
        "EUR",
        row_map,
        "equity_injection",
        styles,
        lambda c, idx=None: f"=IF({idx}=0,{assumptions_map['financing.equity_contribution']},0)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Equity Buybacks / Exit Payouts",
        "EUR",
        row_map,
        "equity_buyback",
        styles,
        lambda c: "=0",
    )
    row = _write_formula_row(
        ws,
        row,
        "Equity at End of Year",
        "EUR",
        row_map,
        "equity_end",
        styles,
        lambda c: f"=C{row_map['equity_start']}+C{row_map['net_income']}+C{row_map['equity_injection']}-C{row_map['dividends']}-C{row_map['equity_buyback']}",
    )

    row = _write_section_label(ws, row + 1, "Check", styles)
    row = _write_formula_row(
        ws,
        row,
        "Total Assets",
        "EUR",
        row_map,
        "total_assets_check",
        styles,
        lambda c: f"=C{row_map['total_assets']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Liabilities + Equity",
        "EUR",
        row_map,
        "total_liabilities_equity",
        styles,
        lambda c: f"=C{row_map['total_liabilities']}+C{row_map['equity_end']}",
    )

    for key in ["total_assets", "total_liabilities", "equity_end", "total_liabilities_equity"]:
        _apply_total_style(ws, row_map[key], styles, 7)

    return {
        "financial_debt": _row_cells("Balance Sheet", row_map["financial_debt"]),
        "cash": _row_cells("Balance Sheet", row_map["cash"]),
        "equity_end": _row_cells("Balance Sheet", row_map["equity_end"]),
        "total_assets": _row_cells("Balance Sheet", row_map["total_assets"]),
    }


def _build_equity_sheet(
    ws,
    assumptions_map: Dict[str, List[str] | str],
    pnl_map: Dict[str, List[str]],
    balance_map: Dict[str, List[str]],
    styles: Dict[str, object],
) -> Dict[str, List[str]]:
    _set_column_widths(ws, [36, 16, 16, 16, 16, 16, 16])
    ws.freeze_panes = "C6"
    _write_title(ws, "Equity Case", 1, 7, styles)

    row = 4
    row = _write_header_row(ws, row, styles)
    row_map: Dict[str, int] = {}

    row = _write_section_label(ws, row, "Capital at Risk (Entry View)", styles)
    row = _write_formula_row(
        ws,
        row,
        "Purchase Price",
        "EUR",
        row_map,
        "purchase_price",
        styles,
        lambda c: f"={assumptions_map['financing.purchase_price']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Debt at Close",
        "EUR",
        row_map,
        "debt_at_close",
        styles,
        lambda c: f"={assumptions_map['financing.senior_debt']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Management Equity",
        "EUR",
        row_map,
        "management_equity",
        styles,
        lambda c: f"={assumptions_map['financing.equity_contribution']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Equity Needed",
        "EUR",
        row_map,
        "total_equity_needed",
        styles,
        lambda c: f"=MAX(C{row_map['purchase_price']}-C{row_map['debt_at_close']},0)",
    )
    row = _write_formula_row(
        ws,
        row,
        "External Investor Equity",
        "EUR",
        row_map,
        "external_equity",
        styles,
        lambda c: f"=MAX(C{row_map['total_equity_needed']}-C{row_map['management_equity']},0)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Equity",
        "EUR",
        row_map,
        "total_equity",
        styles,
        lambda c: f"=MAX(C{row_map['total_equity_needed']},C{row_map['management_equity']},0)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Management Ownership",
        "%",
        row_map,
        "management_share",
        styles,
        lambda c: f"=IF(C{row_map['total_equity']}=0,0,(C{row_map['management_equity']}/C{row_map['total_equity']})*100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "External Ownership",
        "%",
        row_map,
        "external_share",
        styles,
        lambda c: f"=IF(C{row_map['total_equity']}=0,0,(C{row_map['external_equity']}/C{row_map['total_equity']})*100)",
    )

    row = _write_section_label(ws, row + 1, "Exit Value", styles)
    row = _write_formula_row(
        ws,
        row,
        "Exit Multiple",
        "x",
        row_map,
        "exit_multiple",
        styles,
        lambda c: f"={assumptions_map['valuation.multiple']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Final Year EBIT",
        "EUR",
        row_map,
        "final_ebit",
        styles,
        lambda c: f"=INDEX({pnl_map['ebit'][0]}:{pnl_map['ebit'][-1]},5)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Enterprise Value",
        "EUR",
        row_map,
        "enterprise_value",
        styles,
        lambda c: f"=C{row_map['final_ebit']}*C{row_map['exit_multiple']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Net Debt at Exit",
        "EUR",
        row_map,
        "net_debt_exit",
        styles,
        lambda c: f"=INDEX({balance_map['financial_debt'][0]}:{balance_map['financial_debt'][-1]},5)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Excess Cash at Exit",
        "EUR",
        row_map,
        "excess_cash_exit",
        styles,
        lambda c: f"=INDEX({balance_map['cash'][0]}:{balance_map['cash'][-1]},5)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Total Equity Value at Exit",
        "EUR",
        row_map,
        "exit_value",
        styles,
        lambda c: f"=C{row_map['enterprise_value']}-C{row_map['net_debt_exit']}+C{row_map['excess_cash_exit']}",
    )

    row = _write_section_label(ws, row + 1, "Equity Cashflows", styles)
    row = _write_formula_row(
        ws,
        row,
        "Total Equity Cashflow",
        "EUR",
        row_map,
        "equity_cashflow",
        styles,
        lambda c, idx=None: f"=IF({idx}=0,-C{row_map['total_equity']},IF({idx}=4,C{row_map['exit_value']},0))",
    )
    row = _write_formula_row(
        ws,
        row,
        "External Investor Cashflow",
        "EUR",
        row_map,
        "external_cashflow",
        styles,
        lambda c: f"=C{row_map['equity_cashflow']}*(C{row_map['external_share']}/100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Management Cashflow",
        "EUR",
        row_map,
        "management_cashflow",
        styles,
        lambda c: f"=C{row_map['equity_cashflow']}*(C{row_map['management_share']}/100)",
    )

    _apply_total_style(ws, row_map["exit_value"], styles, 7)

    return {
        "exit_value": _row_cells("Equity Case", row_map["exit_value"]),
        "enterprise_value": _row_cells("Equity Case", row_map["enterprise_value"]),
        "equity_cashflow": _row_cells("Equity Case", row_map["equity_cashflow"]),
        "external_share": _row_cells("Equity Case", row_map["external_share"]),
        "management_share": _row_cells("Equity Case", row_map["management_share"]),
        "external_cashflow": _row_cells("Equity Case", row_map["external_cashflow"]),
        "management_cashflow": _row_cells("Equity Case", row_map["management_cashflow"]),
    }


def _build_valuation_sheet(
    ws,
    assumptions_map: Dict[str, List[str] | str],
    pnl_map: Dict[str, List[str]],
    cashflow_map: Dict[str, List[str]],
    balance_map: Dict[str, List[str]],
    equity_map: Dict[str, List[str]],
    styles: Dict[str, object],
) -> None:
    _set_column_widths(ws, [36, 16, 16, 16, 16, 16, 16])
    ws.freeze_panes = "C6"
    _write_title(ws, "Valuation & Purchase Price", 1, 7, styles)

    row = 4
    row = _write_header_row(ws, row, styles)
    row_map: Dict[str, int] = {}

    row = _write_section_label(ws, row, "Seller vs Buyer View", styles)
    row = _write_formula_row(
        ws,
        row,
        "Seller Equity Value",
        "EUR",
        row_map,
        "seller_value",
        styles,
        lambda c: f"=C{row_map['enterprise_value']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Buyer Affordability",
        "EUR",
        row_map,
        "buyer_value",
        styles,
        lambda c: f"=C{row_map['exit_value']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Valuation Gap",
        "EUR",
        row_map,
        "valuation_gap",
        styles,
        lambda c: f"=C{row_map['buyer_value']}-C{row_map['seller_value']}",
    )

    row = _write_section_label(ws, row + 1, "Buyer View (Cash-Based)", styles)
    row = _write_formula_row(
        ws,
        row,
        "Free Cashflow",
        "EUR",
        row_map,
        "free_cashflow",
        styles,
        lambda c, idx=None: f"={cashflow_map['free_cashflow'][idx]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Discount Factor",
        "",
        row_map,
        "discount_factor",
        styles,
        lambda c, idx=None: f"=1/(1+({assumptions_map['valuation.discount_rate']}/100))^{idx+1}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Present Value of FCF",
        "EUR",
        row_map,
        "pv_fcf",
        styles,
        lambda c: f"=C{row_map['free_cashflow']}*C{row_map['discount_factor']}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Cumulative PV of FCF",
        "EUR",
        row_map,
        "cumulative_pv",
        styles,
        lambda c, idx=None: f"=IF({idx}=0,C{row_map['pv_fcf']},INDEX(C{row_map['cumulative_pv']}:G{row_map['cumulative_pv']},{idx})+C{row_map['pv_fcf']})",
    )
    row = _write_formula_row(
        ws,
        row,
        "Net Debt at Close",
        "EUR",
        row_map,
        "net_debt_close",
        styles,
        lambda c: f"={balance_map['financial_debt'][0]}-{balance_map['cash'][0]}",
    )
    row = _write_formula_row(
        ws,
        row,
        "Transaction Costs",
        "EUR",
        row_map,
        "transaction_costs",
        styles,
        lambda c: f"=C{row_map['seller_value']}*({assumptions_map['valuation.txn_cost_pct']}/100)",
    )
    row = _write_formula_row(
        ws,
        row,
        "Equity Value (Buyer View)",
        "EUR",
        row_map,
        "equity_value_buyer",
        styles,
        lambda c: f"=C{row_map['buyer_value']}",
    )

    _apply_total_style(ws, row_map["equity_value_buyer"], styles, 7)


def _build_overview_sheet(
    ws,
    assumptions_map: Dict[str, List[str] | str],
    pnl_map: Dict[str, List[str]],
    equity_map: Dict[str, List[str]],
    revenue_map: Dict[str, List[str]],
    styles: Dict[str, object],
) -> None:
    _set_column_widths(ws, [36, 18, 18, 18, 18, 18, 18])
    ws.freeze_panes = "A6"
    _write_title(ws, "Overview", 1, 7, styles)

    row = 4
    ws.cell(row=row, column=1, value="Deal Snapshot").font = styles["section"]
    row += 1
    ws.cell(row=row, column=1, value="Purchase Price").font = styles["label"]
    ws.cell(row=row, column=2, value=f"={assumptions_map['financing.purchase_price']}").number_format = _currency_format()
    row += 1
    ws.cell(row=row, column=1, value="Debt at Close").font = styles["label"]
    ws.cell(row=row, column=2, value=f"={assumptions_map['financing.senior_debt']}").number_format = _currency_format()
    row += 1
    ws.cell(row=row, column=1, value="Equity at Close").font = styles["label"]
    ws.cell(row=row, column=2, value=f"={assumptions_map['financing.equity_contribution']}").number_format = _currency_format()
    row += 1
    ws.cell(row=row, column=1, value="Debt / EBITDA (Year 0)").font = styles["label"]
    ws.cell(row=row, column=2, value=f"={assumptions_map['financing.senior_debt']}/INDEX({pnl_map['ebitda'][0]}:{pnl_map['ebitda'][-1]},1)").number_format = '0.00"x"'
    row += 1
    ws.cell(row=row, column=1, value="Exit Multiple").font = styles["label"]
    ws.cell(row=row, column=2, value=f"={assumptions_map['valuation.multiple']}").number_format = '0.00"x"'
    row += 2
    ws.cell(row=row, column=1, value="Headline Outcomes (Year 4)").font = styles["section"]
    row += 1
    ws.cell(row=row, column=1, value="Revenue (Year 4)").font = styles["label"]
    ws.cell(row=row, column=2, value=f"=INDEX({revenue_map['final_total_revenue'][0]}:{revenue_map['final_total_revenue'][-1]},5)").number_format = _currency_format()
    row += 1
    ws.cell(row=row, column=1, value="EBITDA (Year 4)").font = styles["label"]
    ws.cell(row=row, column=2, value=f"=INDEX({pnl_map['ebitda'][0]}:{pnl_map['ebitda'][-1]},5)").number_format = _currency_format()
    row += 1
    ws.cell(row=row, column=1, value="Net Income (Year 4)").font = styles["label"]
    ws.cell(row=row, column=2, value=f"=INDEX({pnl_map['net_income'][0]}:{pnl_map['net_income'][-1]},5)").number_format = _currency_format()
    row += 1
    ws.cell(row=row, column=1, value="Equity Value at Exit").font = styles["label"]
    ws.cell(row=row, column=2, value=f"=INDEX({equity_map['exit_value'][0]}:{equity_map['exit_value'][-1]},5)").number_format = _currency_format()


def _write_title(ws, title: str, row: int, last_col: int, styles: Dict[str, object]) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = styles["title"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)


def _write_header_row(ws, row: int, styles: Dict[str, object]) -> int:
    headers = ["Line Item", "Unit"] + YEAR_LABELS
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["header"]
        cell.fill = styles["fill_header"]
        if col == 3:
            cell.fill = styles["fill_year0"]
        cell.alignment = styles["align_center"]
        cell.border = styles["border"]
    return row + 1


def _write_section_label(ws, row: int, title: str, styles: Dict[str, object]) -> int:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = styles["section"]
    cell.fill = styles["fill_section"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    return row + 1


def _write_year_section(
    ws,
    start_row: int,
    title: str,
    rows: List[tuple],
    styles: Dict[str, object],
    mapping: Dict[str, List[str] | str],
) -> int:
    row = _write_section_label(ws, start_row, title, styles)
    row = _write_header_row(ws, row, styles)
    for label, unit, values, key in rows:
        cell_row = row
        ws.cell(row=cell_row, column=1, value=label).font = styles["label"]
        ws.cell(row=cell_row, column=2, value=unit).font = styles["label"]
        mapping[key] = []
        for idx, value in enumerate(values):
            value_to_write = value
            if unit in {"%", "% p.a."} and isinstance(value, (int, float)):
                value_to_write = value * 100
            cell = ws.cell(row=cell_row, column=3 + idx, value=value_to_write)
            cell.font = styles["input"]
            cell.number_format = _input_format_for_unit(unit)
            if idx == 0:
                cell.fill = styles["fill_year0"]
            mapping[key].append(_sheet_ref(ws.title, cell.coordinate))
        row += 1
    return row + 1


def _write_value_section(
    ws,
    start_row: int,
    title: str,
    rows: List[tuple],
    styles: Dict[str, object],
    mapping: Dict[str, List[str] | str],
) -> int:
    row = _write_section_label(ws, start_row, title, styles)
    headers = ["Parameter", "Unit", "Value"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["header"]
        cell.fill = styles["fill_header"]
        cell.alignment = styles["align_center"]
        cell.border = styles["border"]
    row += 1
    for label, unit, value, key in rows:
        ws.cell(row=row, column=1, value=label).font = styles["label"]
        ws.cell(row=row, column=2, value=unit).font = styles["label"]
        value_to_write = value
        if unit in {"%", "% p.a."} and isinstance(value, (int, float)):
            value_to_write = value * 100
        cell = ws.cell(row=row, column=3, value=value_to_write)
        cell.font = styles["input"]
        cell.number_format = _input_format_for_unit(unit)
        mapping[key] = _sheet_ref(ws.title, cell.coordinate)
        row += 1
    return row + 1


def _write_row_from_assumptions(
    ws,
    row: int,
    label: str,
    unit: str,
    assumption_cells: List[str] | str,
    styles: Dict[str, object],
    row_map: Dict[str, int],
    key: str,
) -> int:
    ws.cell(row=row, column=1, value=label).font = styles["label"]
    ws.cell(row=row, column=2, value=unit).font = styles["label"]
    for idx in range(5):
        col = 3 + idx
        cell = ws.cell(row=row, column=col)
        if isinstance(assumption_cells, list):
            cell.value = f"={assumption_cells[idx]}"
        else:
            cell.value = f"={assumption_cells}"
        cell.font = styles["output"]
        cell.number_format = _format_for_unit(unit)
        cell.alignment = styles["align_right"]
        if idx == 0:
            cell.fill = styles["fill_year0"]
    row_map[key] = row
    return row + 1


def _write_formula_row(
    ws,
    row: int,
    label: str,
    unit: str,
    row_map: Dict[str, int],
    key: str,
    styles: Dict[str, object],
    formula_builder,
) -> int:
    ws.cell(row=row, column=1, value=label).font = styles["label"]
    ws.cell(row=row, column=2, value=unit).font = styles["label"]
    for idx in range(5):
        col = 3 + idx
        cell = ws.cell(row=row, column=col)
        if formula_builder.__code__.co_argcount >= 2:
            formula = formula_builder(get_column_letter(col), idx=idx)
        else:
            formula = formula_builder(get_column_letter(col))
        cell.value = formula
        cell.font = styles["output"]
        cell.number_format = _format_for_unit(unit)
        cell.alignment = styles["align_right"]
        if idx == 0:
            cell.fill = styles["fill_year0"]
    row_map[key] = row
    return row + 1


def _apply_total_style(ws, row: int, styles: Dict[str, object], last_col: int) -> None:
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = styles["label_bold"]
        cell.fill = styles["fill_total"]
        if col >= 3:
            cell.number_format = _currency_format()


def _row_cells(sheet_name: str, row: int) -> List[str]:
    return [
        _sheet_ref(sheet_name, f"{get_column_letter(3 + idx)}{row}")
        for idx in range(5)
    ]


def _sheet_ref(sheet_name: str, cell_ref: str) -> str:
    return f"'{sheet_name}'!{cell_ref}"


def _set_column_widths(ws, widths: List[int]) -> None:
    ws.sheet_view.showGridLines = False
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _format_for_unit(unit: str) -> str:
    unit = unit.strip()
    if unit in {"%", "% p.a."}:
        return "0.0"
    if unit in {"x"}:
        return '0.00"x"'
    if unit in {"Years", "Year"}:
        return "0"
    if unit in {"FTE", "Days"}:
        return "0.0"
    if unit == "EUR / %":
        return "0.00"
    if "EUR" in unit:
        return _currency_format()
    return "@"


def _input_format_for_unit(unit: str) -> str:
    unit = unit.strip()
    if unit in {"%", "% p.a."}:
        return "0.0"
    if unit in {"x"}:
        return '0.00"x"'
    if unit in {"Years", "Year"}:
        return "0"
    if unit in {"FTE", "Days"}:
        return "0.0"
    if unit == "EUR / %":
        return "0.00"
    if "EUR" in unit:
        return "#,##0"
    return "@"


def _currency_format() -> str:
    return '#,##0," k€";[Red]-#,##0," k€"'
